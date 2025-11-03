import openpyxl

# Load workbook
file_path = "Portfolio_Syskomp_pA_neu.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

row_num = 37

print(f"Zeile {row_num}:")
print(f"Spalte I (Index 9): '{ws.cell(row=row_num, column=9).value}'")
print(f"Spalte L (Index 12): '{ws.cell(row=row_num, column=12).value}'")

print("\nAlle Zellen in Zeile 37:")
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=row_num, column=col_idx)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"  {col_letter}: {cell.value}")
