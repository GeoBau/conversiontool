"""
Konvertiert Portfolio_Syskomp_pA.xlsx zu CSV
Erstellt automatisch ein Backup der Original-Datei
"""

from openpyxl import load_workbook
import csv
import os
import shutil
from datetime import datetime

def xlsx_to_csv(xlsx_path, csv_path):
    """Konvertiert XLSX zu CSV mit Semikolon als Trennzeichen"""

    print(f"Lade XLSX: {xlsx_path}")

    # Prüfen ob Datei existiert
    if not os.path.exists(xlsx_path):
        print(f"FEHLER: Datei nicht gefunden: {xlsx_path}")
        return False

    # Backup erstellen
    backup_path = xlsx_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy(xlsx_path, backup_path)
    print(f"Backup erstellt: {backup_path}")

    # XLSX laden
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # In CSV schreiben
    print(f"Schreibe CSV: {csv_path}")
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        csv_writer = csv.writer(csvfile, delimiter=';', quoting=csv.QUOTE_MINIMAL)

        row_count = 0
        for row in ws.iter_rows(values_only=True):
            # Konvertiere None zu leerem String
            clean_row = [str(cell).strip() if cell is not None and str(cell).strip() != 'None' else '' for cell in row]
            csv_writer.writerow(clean_row)
            row_count += 1

    print(f"[OK] Erfolgreich {row_count} Zeilen konvertiert")
    print(f"[OK] CSV-Datei: {csv_path}")

    return True

if __name__ == '__main__':
    base_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_file = os.path.join(base_dir, 'Portfolio_Syskomp_pA.xlsx')
    csv_file = os.path.join(base_dir, 'Portfolio_Syskomp_pA.csv')

    success = xlsx_to_csv(xlsx_file, csv_file)

    if success:
        print("\n[OK] Konvertierung abgeschlossen!")
        print(f"  XLSX (Original): {xlsx_file}")
        print(f"  CSV (Neu):       {csv_file}")
    else:
        print("\n[FEHLER] Konvertierung fehlgeschlagen!")
