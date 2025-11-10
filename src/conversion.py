import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
from openpyxl import load_workbook
import os
from collections import defaultdict
from PIL import Image, ImageTk

class ConversionTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Artikelnummern Conversion")
        self.root.geometry("950x580")

        # Data storage
        self.data = defaultdict(dict)  # {column: {value: row_dict}}
        self.column_names = {
            'A': 'Syskomp neue Nummer',
            'B': 'Syskomp alte Nummer',
            'C': 'Syskomp Beschreibung',
            'D': 'Item',
            'E': 'Bosch',
            'F': 'Alvaris Artnr',
            'G': 'Alvaris Matnr',
            'H': 'ASK'
        }

        # Request counter for rate limiting
        self.request_count = 0

        # Compact header: Mode + File in one row
        header_frame = ttk.Frame(root)
        header_frame.pack(fill="x", padx=5, pady=3)

        # Mode selection (left side)
        mode_group = ttk.LabelFrame(header_frame, text="Modus", padding=3)
        mode_group.pack(side="left", padx=2)

        self.mode_var = tk.StringVar(value="intern")
        ttk.Radiobutton(mode_group, text="Intern", variable=self.mode_var, value="intern").pack(side="left", padx=3)
        ttk.Radiobutton(mode_group, text="Extern", variable=self.mode_var, value="extern").pack(side="left", padx=3)

        # Data file (right side)
        file_group = ttk.LabelFrame(header_frame, text="Datenbasis", padding=3)
        file_group.pack(side="left", fill="x", expand=True, padx=2)

        self.file_entry = ttk.Entry(file_group, width=35)
        self.file_entry.pack(side="left", fill="x", expand=True, padx=2)
        self.file_entry.insert(0, "Portfolio_Syskomp_pA.xlsx")

        ttk.Button(file_group, text="...", command=self.browse_file, width=3).pack(side="left", padx=1)
        ttk.Button(file_group, text="Laden", command=self.load_data, width=6).pack(side="left", padx=1)

        self.load_status_label = ttk.Label(file_group, text="", foreground="gray", font=("", 8))
        self.load_status_label.pack(side="left", padx=3)

        # Single conversion - compact
        single_frame = ttk.LabelFrame(root, text="Conversion", padding=5)
        single_frame.pack(fill="both", expand=True, padx=5, pady=3)

        # Selection + Input in one row
        top_row = ttk.Frame(single_frame)
        top_row.pack(fill="x", pady=2)

        ttk.Label(top_row, text="Von:", font=("", 8)).pack(side="left", padx=2)
        self.from_var = tk.StringVar(value="D")
        self.from_combo = ttk.Combobox(top_row, textvariable=self.from_var, width=20, state="readonly", font=("", 8))
        self.from_combo['values'] = [f"{k}: {v}" for k, v in self.column_names.items()]
        self.from_combo.current(3)  # Default: D (Item)
        self.from_combo.pack(side="left", padx=2)

        ttk.Label(top_row, text="→", font=("", 8)).pack(side="left", padx=2)
        self.to_var = tk.StringVar(value="A")
        self.to_combo = ttk.Combobox(top_row, textvariable=self.to_var, width=20, state="readonly", font=("", 8))
        self.to_combo['values'] = [f"{k}: {v}" for k, v in self.column_names.items()]
        self.to_combo.current(0)  # Default: A (Syskomp neu)
        self.to_combo.pack(side="left", padx=2)

        self.input_entry = ttk.Entry(top_row, width=20, font=("", 9))
        self.input_entry.pack(side="left", padx=2, fill="x", expand=True)

        ttk.Button(top_row, text="Konvertieren", command=self.convert_single, width=12).pack(side="left", padx=2)

        # Result + Image side by side
        result_container = ttk.Frame(single_frame)
        result_container.pack(fill="both", expand=True, pady=2)

        # Text result (left)
        text_frame = ttk.Frame(result_container)
        text_frame.pack(side="left", fill="both", expand=True, padx=(0, 2))

        self.result_text = scrolledtext.ScrolledText(text_frame, height=10, wrap="word", font=("", 9))
        self.result_text.pack(fill="both", expand=True)

        # Image display (right)
        img_frame = ttk.Frame(result_container, relief="sunken", borderwidth=1)
        img_frame.pack(side="left", fill="both", padx=(2, 0))

        self.image_label = ttk.Label(img_frame, text="", anchor="center")
        self.image_label.pack(fill="both", expand=True)
        self.current_image = None  # Keep reference to prevent garbage collection

        # Batch conversion - compact
        batch_frame = ttk.LabelFrame(root, text="Batch", padding=5)
        batch_frame.pack(fill="both", expand=True, padx=5, pady=3)

        batch_controls = ttk.Frame(batch_frame)
        batch_controls.pack(fill="x", pady=2)

        ttk.Button(batch_controls, text="Laden", command=self.load_batch_file, width=8).pack(side="left", padx=2)
        ttk.Button(batch_controls, text="Verarbeiten", command=self.process_batch, width=10).pack(side="left", padx=2)
        ttk.Button(batch_controls, text="Speichern", command=self.save_batch_result, width=10).pack(side="left", padx=2)

        ttk.Label(batch_controls, text="→", font=("", 8)).pack(side="left", padx=5)
        self.batch_target_var = tk.StringVar(value="A")
        batch_target_combo = ttk.Combobox(batch_controls, textvariable=self.batch_target_var, width=18, state="readonly", font=("", 8))
        batch_target_combo['values'] = ["A: Syskomp neue Nummer", "B: Syskomp alte Nummer"]
        batch_target_combo.current(0)
        batch_target_combo.pack(side="left", padx=2)

        # Batch result
        self.batch_text = scrolledtext.ScrolledText(batch_frame, height=8, wrap="word", font=("", 9))
        self.batch_text.pack(fill="both", expand=True, pady=2)

        # Status bar
        self.status_label = ttk.Label(root, text="Bereit", relief="sunken", font=("", 8))
        self.status_label.pack(side="bottom", fill="x", padx=2, pady=1)

        # Auto-load data
        self.load_data()

    def browse_file(self):
        filepath = filedialog.askopenfilename(
            title="Excel-Datei auswählen",
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")]
        )
        if filepath:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, filepath)

    def load_data(self):
        filepath = self.file_entry.get().strip()

        if not filepath or not os.path.exists(filepath):
            self.load_status_label.config(text="❌ Datei nicht gefunden", foreground="red")
            return

        try:
            self.load_status_label.config(text="⏳ Lade Daten...", foreground="blue")
            self.root.update()

            wb = load_workbook(filepath, data_only=True)
            ws = wb.active

            # Clear existing data
            self.data.clear()

            # Load data (skip header row)
            for row_idx in range(2, ws.max_row + 1):
                row_dict = {}

                # Read columns A-H
                for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'], start=1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_dict[col_letter] = str(cell_value).strip() if cell_value else ""

                # Index by each column (for fast lookup)
                for col_letter in ['A', 'B', 'D', 'E', 'F', 'H']:  # Searchable columns
                    value = row_dict.get(col_letter, "")
                    if value:
                        self.data[col_letter][value] = row_dict

            total_rows = ws.max_row - 1
            self.load_status_label.config(text=f"✓ {total_rows} Zeilen geladen", foreground="green")
            self.status_label.config(text=f"Daten geladen: {total_rows} Zeilen")

        except Exception as e:
            self.load_status_label.config(text=f"❌ Fehler: {e}", foreground="red")

    def get_column_letter(self, selection):
        """Extract column letter from 'X: Description' format"""
        return selection.split(':')[0].strip()

    def load_and_display_image(self, artnr, source_type):
        """Load and display image based on article number and source type
        source_type: 'bosch', 'item', 'alvaris'
        """
        # Clear previous image
        self.image_label.config(image='', text='')
        self.current_image = None

        if not artnr or artnr == '-':
            return

        # Determine image directory
        script_dir = os.path.dirname(os.path.abspath(__file__))

        # Try multiple possible directory names
        possible_dirs = []
        if source_type == 'bosch':
            possible_dirs = [
                os.path.join(script_dir, "ASK_CATALOG", "ask-bosch-images"),
                os.path.join(script_dir, "ASK_CATALOG", "ASKbosch-all-images"),
            ]
        elif source_type == 'item':
            possible_dirs = [
                os.path.join(script_dir, "ASK_CATALOG", "ask-item-images"),
                os.path.join(script_dir, "ASK_CATALOG", "ASKitem-all-images"),
            ]
        elif source_type == 'alvaris':
            possible_dirs = [
                os.path.join(script_dir, "ALVARIS_CATALOG", "alvaris-images"),
                os.path.join(script_dir, "ALVARIS_CATALOG", "alvaris-all-images"),
            ]
        else:
            return

        # Try to find image in any of the possible directories
        img_path = None
        for img_dir in possible_dirs:
            test_path = os.path.join(img_dir, f"{artnr}.png")
            if os.path.exists(test_path):
                img_path = test_path
                break

        if not img_path:
            self.image_label.config(text=f"⚠️ Bild nicht gefunden: {artnr}.png")
            return

        try:
            # Load image
            img = Image.open(img_path)

            # For Alvaris: crop to upper 70%
            if source_type == 'alvaris':
                width, height = img.size
                crop_height = int(height * 0.7)
                img = img.crop((0, 0, width, crop_height))

            # Resize to fit display (max 300px wide, maintain aspect ratio)
            max_width = 300
            max_height = 250
            width, height = img.size

            # Scale to fit within bounds
            ratio = min(max_width / width, max_height / height)
            if ratio < 1:
                new_width = int(width * ratio)
                new_height = int(height * ratio)
                img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)

            # Convert to PhotoImage
            photo = ImageTk.PhotoImage(img)
            self.current_image = photo  # Keep reference

            # Display
            self.image_label.config(image=photo, text='')

        except Exception as e:
            self.image_label.config(text=f"⚠️ Fehler beim Laden: {e}")

    def validate_conversion(self, from_col, to_col, mode):
        """Validate if conversion is allowed based on rules"""
        # A or B must be involved
        if from_col not in ['A', 'B'] and to_col not in ['A', 'B']:
            return False, "❌ Mindestens A oder B muss involviert sein!"

        # Extern: Only allow X → A or X → B
        if mode == "extern":
            if to_col not in ['A', 'B']:
                return False, "❌ Extern: Nur Conversion zu Syskomp A oder B erlaubt!"

        return True, ""

    def convert_single(self):
        from_col = self.get_column_letter(self.from_var.get())
        to_col = self.get_column_letter(self.to_var.get())
        search_value = self.input_entry.get().strip()
        mode = self.mode_var.get()

        self.result_text.delete(1.0, tk.END)

        if not search_value:
            self.result_text.insert(tk.END, "⚠️ Bitte Suchnummer eingeben")
            return

        if not self.data:
            self.result_text.insert(tk.END, "⚠️ Bitte zuerst Daten laden")
            return

        # Validate conversion
        valid, error_msg = self.validate_conversion(from_col, to_col, mode)
        if not valid:
            self.result_text.insert(tk.END, error_msg)
            return

        # Check rate limiting for extern
        if mode == "extern":
            self.request_count += 1
            if self.request_count >= 50:
                self.result_text.insert(tk.END, f"⚠️ Rate Limit: Anfrage #{self.request_count}\n")
                self.result_text.insert(tk.END, "Externe Anfragen werden ab der 50. Anfrage verlangsamt.\n\n")
                import time
                time.sleep(2)  # Slow down

        # Search in data
        row_data = self.data.get(from_col, {}).get(search_value)

        if not row_data:
            self.result_text.insert(tk.END, f"❌ '{search_value}' nicht gefunden in Spalte {from_col}\n")
            self.status_label.config(text="Nicht gefunden")
            return

        # Get result
        if to_col in ['F', 'G']:  # Alvaris special case
            artnr = row_data.get('F', '-')
            matnr = row_data.get('G', '-')
            result = f"{artnr} / {matnr}"
        else:
            result = row_data.get(to_col, '-')

        # Get description
        description = row_data.get('C', '')
        description = description.replace(';', '\n')

        # Display result
        self.result_text.insert(tk.END, f"✓ Conversion erfolgreich:\n\n")
        self.result_text.insert(tk.END, f"{search_value} → {result}\n\n")
        self.result_text.insert(tk.END, f"Beschreibung:\n{description}\n")

        self.status_label.config(text=f"Conversion: {from_col} → {to_col}")

        # Display image if available
        # Check if this is an ASK or Alvaris article
        ask_artnr = row_data.get('H', '')  # ASK article number
        alvaris_artnr = row_data.get('F', '')  # Alvaris article number
        bosch_nr = row_data.get('E', '')  # Bosch number
        item_nr = row_data.get('D', '')  # Item number

        if alvaris_artnr and alvaris_artnr != '-':
            # Alvaris article - use Alvaris image
            self.load_and_display_image(alvaris_artnr, 'alvaris')
        elif ask_artnr and ask_artnr != '-':
            # ASK article - determine if Bosch or Item
            # Check which type of number exists
            if bosch_nr and len(bosch_nr) == 10:
                # Bosch number (10 digits)
                self.load_and_display_image(ask_artnr, 'bosch')
            elif item_nr and '.' in item_nr:
                # Item number (x.x.x.x format)
                self.load_and_display_image(ask_artnr, 'item')
            else:
                # Default to ask_artnr
                self.load_and_display_image(ask_artnr, 'item')

    def load_batch_file(self):
        filepath = filedialog.askopenfilename(
            title="Batch-Datei laden",
            filetypes=[("Text/CSV Dateien", "*.txt;*.csv"), ("Alle Dateien", "*.*")]
        )

        if filepath:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()

                self.batch_text.delete(1.0, tk.END)
                self.batch_text.insert(tk.END, content)
                self.status_label.config(text=f"Batch-Datei geladen: {os.path.basename(filepath)}")

            except Exception as e:
                messagebox.showerror("Fehler", f"Fehler beim Laden: {e}")

    def process_batch(self):
        # Get input numbers (one per line)
        input_text = self.batch_text.get(1.0, tk.END).strip()
        if not input_text:
            messagebox.showwarning("Warnung", "Keine Daten zum Verarbeiten")
            return

        if not self.data:
            messagebox.showerror("Fehler", "Bitte zuerst Daten laden")
            return

        lines = [line.strip() for line in input_text.split('\n') if line.strip()]
        target_col = self.batch_target_var.get().split(':')[0].strip()

        self.batch_text.delete(1.0, tk.END)
        self.batch_text.insert(tk.END, "Batch Conversion Ergebnisse:\n")
        self.batch_text.insert(tk.END, "=" * 60 + "\n\n")

        found_count = 0
        not_found_count = 0

        for search_value in lines:
            # Try to find in any searchable column
            row_data = None
            found_col = None

            for col in ['A', 'B', 'D', 'E', 'F', 'H']:
                if search_value in self.data.get(col, {}):
                    row_data = self.data[col][search_value]
                    found_col = col
                    break

            if row_data:
                result = row_data.get(target_col, '-')
                description = row_data.get('C', '').replace(';', ' | ')

                self.batch_text.insert(tk.END, f"✓ {search_value} → {result}\n")
                if description:
                    self.batch_text.insert(tk.END, f"  {description}\n")
                self.batch_text.insert(tk.END, "\n")
                found_count += 1
            else:
                self.batch_text.insert(tk.END, f"❌ {search_value} → Nicht gefunden\n\n")
                not_found_count += 1

        self.batch_text.insert(tk.END, "=" * 60 + "\n")
        self.batch_text.insert(tk.END, f"Ergebnis: {found_count} gefunden, {not_found_count} nicht gefunden\n")

        self.status_label.config(text=f"Batch verarbeitet: {found_count}/{len(lines)} gefunden")

    def save_batch_result(self):
        content = self.batch_text.get(1.0, tk.END).strip()

        if not content or content == "":
            messagebox.showwarning("Warnung", "Nichts zum Speichern")
            return

        filepath = filedialog.asksaveasfilename(
            title="Ergebnis speichern",
            defaultextension=".txt",
            filetypes=[("Text Dateien", "*.txt"), ("CSV Dateien", "*.csv"), ("Alle Dateien", "*.*")]
        )

        if filepath:
            try:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(content)

                messagebox.showinfo("Erfolg", f"Ergebnis gespeichert: {os.path.basename(filepath)}")
                self.status_label.config(text=f"Gespeichert: {os.path.basename(filepath)}")

            except Exception as e:
                messagebox.showerror("Fehler", f"Fehler beim Speichern: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ConversionTool(root)
    root.mainloop()
