from playwright.sync_api import sync_playwright
import time
import os
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import threading
import re
from openpyxl import load_workbook

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; Alvaris-Search/1.0)"
}

class AlvarisSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Alvaris Item Number Search")
        self.root.geometry("800x700")
        self.is_searching = False

        # Input Excel File Selection
        input_frame = ttk.LabelFrame(root, text="Eingabe Excel-Datei", padding=10)
        input_frame.pack(fill="x", padx=10, pady=5)

        input_inner = ttk.Frame(input_frame)
        input_inner.pack(fill="x")

        self.input_file_entry = ttk.Entry(input_inner, width=50)
        self.input_file_entry.pack(side="left", fill="x", expand=True)
        self.input_file_entry.insert(0, "Portfolio_Syskomp_pA.xlsx")

        ttk.Button(input_inner, text="Durchsuchen...", command=self.browse_input_file).pack(side="left", padx=5)

        ttk.Label(input_frame, text="💡 Spalte F = Bosch-Nummern für Alvaris-Suche | Zeilen mit Wert in Spalte N werden übersprungen",
                  font=("", 8), foreground="gray").pack(anchor="w", pady=(5, 0))

        # Output Excel File Selection
        output_frame = ttk.LabelFrame(root, text="Ausgabe Excel-Datei", padding=10)
        output_frame.pack(fill="x", padx=10, pady=5)

        output_inner = ttk.Frame(output_frame)
        output_inner.pack(fill="x")

        self.output_file_entry = ttk.Entry(output_inner, width=50)
        self.output_file_entry.pack(side="left", fill="x", expand=True)
        self.output_file_entry.insert(0, "Portfolio_Syskomp_pA.xlsx")

        ttk.Button(output_inner, text="Durchsuchen...", command=self.browse_output_file).pack(side="left", padx=5)

        ttk.Label(output_frame, text="💡 Ergebnisse werden in Spalte N (artnr) und O (matnr) geschrieben",
                  font=("", 8), foreground="gray").pack(anchor="w", pady=(5, 0))

        # Settings
        settings_frame = ttk.LabelFrame(root, text="Einstellungen", padding=10)
        settings_frame.pack(fill="x", padx=10, pady=5)

        settings_inner = ttk.Frame(settings_frame)
        settings_inner.pack(fill="x")

        ttk.Label(settings_inner, text="Pause zw. Suchen (Sek):").pack(side="left")
        self.search_pause_entry = ttk.Entry(settings_inner, width=5)
        self.search_pause_entry.insert(0, "3")
        self.search_pause_entry.pack(side="left", padx=5)

        ttk.Label(settings_inner, text="Startzeile (Excel):").pack(side="left", padx=(20, 5))
        self.start_row_entry = ttk.Entry(settings_inner, width=5)
        self.start_row_entry.insert(0, "2")  # Default: Skip header row
        self.start_row_entry.pack(side="left", padx=5)

        ttk.Label(settings_inner, text="Max. Zeilen (leer = alle):").pack(side="left", padx=(20, 5))
        self.max_rows_entry = ttk.Entry(settings_inner, width=5)
        self.max_rows_entry.pack(side="left", padx=5)

        # Debug mode
        self.debug_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(settings_inner, text="Browser anzeigen", variable=self.debug_var).pack(side="left", padx=(20, 5))

        # Control Buttons
        button_frame = ttk.Frame(root, padding=10)
        button_frame.pack(fill="x", padx=10)

        self.start_btn = ttk.Button(button_frame, text="▶ Suche starten", command=self.start_search)
        self.start_btn.pack(side="left", padx=5)

        self.stop_btn = ttk.Button(button_frame, text="⏹ Stoppen", command=self.stop_search, state="disabled")
        self.stop_btn.pack(side="left", padx=5)

        # Status Display
        status_frame = ttk.Frame(button_frame)
        status_frame.pack(side="left", padx=20, fill="x", expand=True)

        self.status_label = ttk.Label(status_frame, text="Bereit", font=("", 10, "bold"), foreground="blue")
        self.status_label.pack(anchor="w")

        self.current_item_label = ttk.Label(status_frame, text="", font=("", 9), foreground="gray")
        self.current_item_label.pack(anchor="w")

        # Progress Bar
        self.progress = ttk.Progressbar(root, mode="indeterminate")
        self.progress.pack(fill="x", padx=10, pady=5)

        # Log/Status Display
        log_frame = ttk.LabelFrame(root, text="Status Log", padding=10)
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=20, wrap="word", state="normal")
        self.log_text.pack(fill="both", expand=True)

        # Make read-only but allow selection/copy
        def on_key(event):
            if event.state & 0x4:  # Control key
                return
            if event.keysym in ('Left', 'Right', 'Up', 'Down', 'Home', 'End', 'Prior', 'Next'):
                return
            return "break"

        self.log_text.bind("<Key>", on_key)

    def browse_input_file(self):
        filepath = filedialog.askopenfilename(
            title="Excel-Datei auswählen",
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")]
        )
        if filepath:
            self.input_file_entry.delete(0, tk.END)
            self.input_file_entry.insert(0, filepath)

    def browse_output_file(self):
        filepath = filedialog.asksaveasfilename(
            title="Ausgabe-Excel-Datei",
            defaultextension=".xlsx",
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
            confirmoverwrite=True
        )
        if filepath:
            self.output_file_entry.delete(0, tk.END)
            self.output_file_entry.insert(0, filepath)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)

    def update_status(self, current, total, bosch_number=""):
        """Update status display"""
        self.status_label.config(text=f"Fortschritt: {current}/{total}")
        if bosch_number:
            self.current_item_label.config(text=f"Suche Bosch: {bosch_number}")

    def start_search(self):
        input_file = self.input_file_entry.get().strip()
        output_file = self.output_file_entry.get().strip()

        if not input_file:
            messagebox.showerror("Fehler", "Bitte wählen Sie eine Eingabe-Excel-Datei.")
            return

        if not output_file:
            messagebox.showerror("Fehler", "Bitte wählen Sie eine Ausgabe-Excel-Datei.")
            return

        if not os.path.exists(input_file):
            messagebox.showerror("Fehler", f"Eingabe-Datei nicht gefunden: {input_file}")
            return

        try:
            start_row = int(self.start_row_entry.get().strip() or "2")
        except ValueError:
            messagebox.showerror("Fehler", "Ungültige Startzeile.")
            return

        max_rows_str = self.max_rows_entry.get().strip()
        max_rows = None
        if max_rows_str:
            try:
                max_rows = int(max_rows_str)
            except ValueError:
                messagebox.showerror("Fehler", "Ungültige Max. Zeilen.")
                return

        # Start search in thread
        self.is_searching = True
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.progress.start()
        self.log_text.delete(1.0, tk.END)

        thread = threading.Thread(target=self.run_search, args=(input_file, output_file, start_row, max_rows))
        thread.daemon = True
        thread.start()

    def stop_search(self):
        self.is_searching = False
        self.log("⏹ Abbruch angefordert...")

    def run_search(self, input_file, output_file, start_row, max_rows):
        try:
            self.log(f"📂 Lade Excel-Datei: {input_file}")

            # Load workbook
            wb = load_workbook(input_file)
            ws = wb.active

            self.log(f"✓ Excel-Datei geladen: {ws.title}")

            # Get search pause
            try:
                search_pause = float(self.search_pause_entry.get())
            except:
                search_pause = 3.0

            debug_mode = self.debug_var.get()

            # Count rows with data in column F (Bosch numbers) and empty column N
            bosch_numbers = []
            row_mapping = {}  # Map index to actual row number
            skipped_count = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=False), start=start_row):
                if max_rows and (row_idx - start_row + 1) > max_rows:
                    break

                # Column F is index 5 (0-based), Column N is index 13
                bosch_cell = row[5]  # Column F (Bosch number)
                artnr_cell = row[13]  # Column N

                bosch_value = bosch_cell.value
                artnr_value = artnr_cell.value

                # Only process if column F has value AND column N is empty
                if bosch_value and str(bosch_value).strip():
                    if artnr_value and str(artnr_value).strip():
                        # Column N already has value, skip this row
                        skipped_count += 1
                    else:
                        # Column N is empty, add to processing list
                        bosch_numbers.append(str(bosch_value).strip())
                        row_mapping[len(bosch_numbers) - 1] = row_idx

            if not bosch_numbers:
                if skipped_count > 0:
                    self.log(f"⚠️ Keine neuen Bosch-Nummern zu verarbeiten ({skipped_count} bereits verarbeitet)")
                    messagebox.showinfo("Info", f"Alle {skipped_count} Bosch-Nummern wurden bereits verarbeitet.")
                else:
                    self.log("⚠️ Keine Bosch-Nummern in Spalte F gefunden")
                    messagebox.showwarning("Warnung", "Keine Bosch-Nummern gefunden in Spalte F")
                self.finish_search()
                return

            self.log(f"📋 {len(bosch_numbers)} Bosch-Nummern gefunden in Spalte F")
            if skipped_count > 0:
                self.log(f"⏭️ {skipped_count} Zeilen übersprungen (bereits verarbeitet)")
            self.log(f"🔍 Starte Alvaris-Suche...")

            # Scrape Alvaris
            results = self.search_alvaris(bosch_numbers, search_pause, debug_mode)

            if not self.is_searching:
                self.log("❌ Suche abgebrochen.")
                self.finish_search()
                return

            # Write results back to Excel
            self.log(f"\n💾 Schreibe Ergebnisse in Excel...")

            for idx, (bosch_number, artnr, matnr) in enumerate(results):
                if idx in row_mapping:
                    row_num = row_mapping[idx]
                    # Column N = index 13 (0-based), Column O = index 14
                    ws.cell(row=row_num, column=14, value=artnr)  # Column N
                    ws.cell(row=row_num, column=15, value=matnr)  # Column O

                    if artnr != "-":
                        self.log(f"  ✓ Zeile {row_num}: Bosch {bosch_number} → Alvaris {artnr} / {matnr}")
                    else:
                        self.log(f"  ⚠ Zeile {row_num}: Bosch {bosch_number} → Nichts gefunden")

            # Save workbook
            wb.save(output_file)
            self.log(f"\n✅ Fertig! Ergebnisse gespeichert in: {output_file}")
            messagebox.showinfo("Erfolg", f"Suche abgeschlossen!\n{len(results)} Bosch-Nummern verarbeitet.")

        except Exception as e:
            self.log(f"❌ Fehler: {e}")
            messagebox.showerror("Fehler", f"Fehler beim Verarbeiten: {e}")

        finally:
            self.finish_search()

    def search_alvaris(self, bosch_numbers, search_pause, debug_mode):
        """Search Alvaris website for each Bosch number"""
        results = []

        with sync_playwright() as p:
            self.log("🌐 Starte Browser...")
            browser = p.chromium.launch(headless=not debug_mode)
            page = browser.new_page()
            page.set_extra_http_headers(HEADERS)

            # Wait 30 seconds on first start for manual interactions
            self.log("⏸️ Warte 30 Sekunden... (Zeit für Cookie-Banner, Captcha, etc.)")
            for remaining in range(30, 0, -5):
                if not self.is_searching:
                    browser.close()
                    return []
                self.log(f"⏳ Noch {remaining} Sekunden...")
                time.sleep(5)
            self.log("▶️ Starte jetzt mit der Suche...")

            for idx, bosch_number in enumerate(bosch_numbers):
                if not self.is_searching:
                    break

                current = idx + 1
                total = len(bosch_numbers)

                self.log(f"\n🔍 [{current}/{total}] Suche Bosch-Nummer: {bosch_number}")
                self.update_status(current, total, bosch_number)

                try:
                    # Build search URL
                    search_url = f"https://www.alvaris.com/de/?s={bosch_number}&trp-form-language=de"

                    # Load page
                    page.goto(search_url, wait_until="networkidle", timeout=30000)
                    time.sleep(2)  # Wait for page to fully render

                    # Check if nothing found
                    nothing_found = page.locator("h1:has-text('Nichts gefunden')").count() > 0

                    if nothing_found:
                        self.log(f"  ⚠️ Nichts gefunden für Bosch {bosch_number}")
                        results.append((bosch_number, "-", "-"))
                    else:
                        # Verify search number in breadcrumb
                        breadcrumb = page.locator('span[aria-current="page"]').first
                        if breadcrumb.count() > 0:
                            breadcrumb_text = breadcrumb.inner_text().strip()
                            if bosch_number not in breadcrumb_text:
                                self.log(f"  ⚠️ Bosch-Nummer nicht in Breadcrumb gefunden")
                                results.append((bosch_number, "-", "-"))
                                continue

                        # Extract article info
                        artnr = "-"
                        matnr = "-"

                        # Try to find article description text
                        article_text = page.locator("article").first
                        if article_text.count() > 0:
                            full_text = article_text.inner_text()

                            # Parse "Artikelbeschreibung Artikelnummer 1010634 / WINAL ..."
                            match = re.search(r'Artikelnummer\s+(\d+)\s*/\s*(\S+)', full_text)
                            if match:
                                artnr = match.group(1)
                                matnr = match.group(2)
                                self.log(f"  ✓ Alvaris gefunden: {artnr} / {matnr}")
                            else:
                                self.log(f"  ⚠️ Artikelnummer-Format nicht gefunden")

                        results.append((bosch_number, artnr, matnr))

                    # Pause before next search
                    if self.is_searching and current < total:
                        time.sleep(search_pause)

                except Exception as e:
                    self.log(f"  ❌ Fehler bei Bosch {bosch_number}: {e}")
                    results.append((bosch_number, "-", "-"))

            browser.close()
            self.log("🌐 Browser geschlossen.")

        return results

    def finish_search(self):
        self.is_searching = False
        self.progress.stop()
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")

if __name__ == "__main__":
    root = tk.Tk()
    app = AlvarisSearchApp(root)
    root.mainloop()
