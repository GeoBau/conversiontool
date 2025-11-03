import openpyxl
import re

# Load workbook
file_path = "Portfolio_Syskomp_pA_neu.xlsx"
print(f"Lade Excel-Datei: {file_path}")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"Worksheet: {ws.title}")
print(f"Max Row: {ws.max_row}")

# Pattern for x.x.x.x (e.g., 0.0.411.14)
pattern = re.compile(r'\d+\.\d+\.\d+\.\d+')

changes_made = 0

# Iterate through all rows
for row_idx in range(1, ws.max_row + 1):
    # Check if column I (index 9) is empty
    cell_i = ws.cell(row=row_idx, column=9)

    if cell_i.value is None or str(cell_i.value).strip() == "":
        # Column I is empty, search for pattern in entire row
        found_pattern = None

        # Search up to column Z (26) to ensure we catch all data
        for col_idx in range(1, 27):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_str = str(cell.value)
                match = pattern.search(cell_str)
                if match:
                    found_pattern = match.group(0)
                    print(f"Zeile {row_idx}: Gefunden '{found_pattern}' in Spalte {openpyxl.utils.get_column_letter(col_idx)}")
                    break

        # If pattern found, write to column I
        if found_pattern:
            ws.cell(row=row_idx, column=9, value=found_pattern)
            changes_made += 1
            print(f"  > Geschrieben in Spalte I")

print(f"\nFertig! {changes_made} Änderungen vorgenommen.")

# Save workbook
wb.save(file_path)
print(f"Datei gespeichert: {file_path}")
