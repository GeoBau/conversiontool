import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import os
from collections import defaultdict
from PIL import Image, ImageTk

# Syskomp Colors
SYSKOMP_GREEN = "#409f95"
BG_LIGHT = "#eeeeee"
BG_WHITE = "#ffffff"
TEXT_DARK = "#333333"
TEXT_GRAY = "#666666"
BORDER_COLOR = "#cccccc"

class CompactText(tk.Text):
    """Compact scrollable text widget"""
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.config(
            font=("Segoe UI", 9),
            wrap="word",
            relief="solid",
            borderwidth=1,
            bg=BG_WHITE,
            fg=TEXT_DARK
        )

class ConversionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("conversionTool")
        self.root.geometry("1000x550")
        self.root.configure(bg=BG_WHITE)

        # Data
        self.data = defaultdict(dict)
        self.column_names = {
            'A': 'Syskomp neu', 'B': 'Syskomp alt', 'C': 'Beschreibung',
            'D': 'Item', 'E': 'Bosch', 'F': 'Alvaris Art',
            'G': 'Alvaris Mat', 'H': 'ASK'
        }
        self.request_count = 0
        self.current_image = None

        # Header
        header = tk.Frame(root, bg=SYSKOMP_GREEN, height=30)
        header.pack(fill="x", padx=4, pady=(4, 0))
        header.pack_propagate(False)

        tk.Label(
            header, text="conversionTool",
            bg=SYSKOMP_GREEN, fg="white",
            font=("Segoe UI", 14, "bold")
        ).pack(side="left", padx=6, pady=2)

        # Mode + File in header
        mode_frame = tk.Frame(header, bg=SYSKOMP_GREEN)
        mode_frame.pack(side="right", padx=6)

        self.mode_var = tk.StringVar(value="intern")
        tk.Radiobutton(
            mode_frame, text="Intern", variable=self.mode_var, value="intern",
            bg=SYSKOMP_GREEN, fg="white", selectcolor=SYSKOMP_GREEN,
            font=("Segoe UI", 8), activebackground=SYSKOMP_GREEN
        ).pack(side="left", padx=2)
        tk.Radiobutton(
            mode_frame, text="Extern", variable=self.mode_var, value="extern",
            bg=SYSKOMP_GREEN, fg="white", selectcolor=SYSKOMP_GREEN,
            font=("Segoe UI", 8), activebackground=SYSKOMP_GREEN
        ).pack(side="left", padx=2)

        # Main container - 2 columns
        main = tk.Frame(root, bg=BG_WHITE)
        main.pack(fill="both", expand=True, padx=4, pady=4)

        # Left column - Single Conversion
        left = tk.Frame(main, bg=BG_LIGHT, relief="flat", bd=1)
        left.pack(side="left", fill="both", expand=True, padx=(0, 2))

        tk.Label(
            left, text="Einzelsuche",
            bg=BG_LIGHT, fg=SYSKOMP_GREEN,
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w", padx=6, pady=(4, 2))

        # Input row
        input_row = tk.Frame(left, bg=BG_LIGHT)
        input_row.pack(fill="x", padx=6, pady=2)

        # From dropdown
        tk.Label(input_row, text="Von:", bg=BG_LIGHT, fg=TEXT_GRAY, font=("Segoe UI", 8)).pack(side="left", padx=(0,2))
        self.from_var = tk.StringVar(value="B")
        from_combo = ttk.Combobox(
            input_row, textvariable=self.from_var,
            width=15, state="readonly", font=("Segoe UI", 8)
        )
        from_combo['values'] = [f"{k}:{v}" for k, v in self.column_names.items()]
        from_combo.current(1)
        from_combo.pack(side="left", padx=2)

        tk.Label(input_row, text="→", bg=BG_LIGHT, fg=TEXT_GRAY, font=("Segoe UI", 8)).pack(side="left", padx=2)

        # To dropdown
        self.to_var = tk.StringVar(value="A")
        to_combo = ttk.Combobox(
            input_row, textvariable=self.to_var,
            width=15, state="readonly", font=("Segoe UI", 8)
        )
        to_combo['values'] = [f"{k}:{v}" for k, v in self.column_names.items()]
        to_combo.current(0)
        to_combo.pack(side="left", padx=2)

        # Search input
        self.search_entry = tk.Entry(
            input_row, font=("Segoe UI", 9),
            relief="solid", bd=1, bg=BG_WHITE
        )
        self.search_entry.pack(side="left", fill="x", expand=True, padx=2)
        self.search_entry.bind('<Return>', lambda e: self.convert_single())

        # Search button
        search_btn = tk.Button(
            input_row, text="Suchen", command=self.convert_single,
            bg=SYSKOMP_GREEN, fg="white", font=("Segoe UI", 8, "bold"),
            relief="flat", bd=0, padx=8, cursor="hand2"
        )
        search_btn.pack(side="left", padx=2)

        # Result area with image
        result_frame = tk.Frame(left, bg=BG_LIGHT)
        result_frame.pack(fill="both", expand=True, padx=6, pady=4)

        # Text result (left)
        self.result_text = CompactText(result_frame, height=10, width=40)
        self.result_text.pack(side="left", fill="both", expand=True, padx=(0, 2))

        # Image (right)
        img_container = tk.Frame(result_frame, bg=BG_WHITE, relief="solid", bd=1)
        img_container.pack(side="left", fill="both")
        self.image_label = tk.Label(img_container, bg=BG_WHITE, text="")
        self.image_label.pack(padx=2, pady=2)

        # Right column - Batch Conversion
        right = tk.Frame(main, bg=BG_LIGHT, relief="flat", bd=1)
        right.pack(side="left", fill="both", expand=True, padx=(2, 0))

        tk.Label(
            right, text="Batch-Konvertierung",
            bg=BG_LIGHT, fg=SYSKOMP_GREEN,
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w", padx=6, pady=(4, 2))

        # File controls
        file_row = tk.Frame(right, bg=BG_LIGHT)
        file_row.pack(fill="x", padx=6, pady=2)

        self.file_label = tk.Label(
            file_row, text="Keine Datei", bg=BG_WHITE,
            fg=TEXT_GRAY, font=("Segoe UI", 8), relief="solid", bd=1,
            anchor="w", padx=4
        )
        self.file_label.pack(side="left", fill="x", expand=True, padx=(0,2))

        tk.Button(
            file_row, text="Wählen", command=self.load_batch_file,
            bg=SYSKOMP_GREEN, fg="white", font=("Segoe UI", 8, "bold"),
            relief="flat", bd=0, cursor="hand2"
        ).pack(side="left", padx=2)

        # Batch options
        opt_row = tk.Frame(right, bg=BG_LIGHT)
        opt_row.pack(fill="x", padx=6, pady=2)

        tk.Label(opt_row, text="→", bg=BG_LIGHT, fg=TEXT_GRAY, font=("Segoe UI", 8)).pack(side="left", padx=2)

        self.batch_target_var = tk.StringVar(value="A")
        target_combo = ttk.Combobox(
            opt_row, textvariable=self.batch_target_var,
            width=15, state="readonly", font=("Segoe UI", 8)
        )
        target_combo['values'] = ["A:Syskomp neu", "B:Syskomp alt"]
        target_combo.current(0)
        target_combo.pack(side="left", padx=2)

        tk.Button(
            opt_row, text="Verarbeiten", command=self.process_batch,
            bg=SYSKOMP_GREEN, fg="white", font=("Segoe UI", 8, "bold"),
            relief="flat", bd=0, cursor="hand2"
        ).pack(side="left", padx=2)

        tk.Button(
            opt_row, text="Speichern", command=self.save_batch,
            bg=SYSKOMP_GREEN, fg="white", font=("Segoe UI", 8, "bold"),
            relief="flat", bd=0, cursor="hand2"
        ).pack(side="left", padx=2)

        # Batch results
        self.batch_text = CompactText(right, height=15)
        self.batch_text.pack(fill="both", expand=True, padx=6, pady=4)

        # Status bar
        status_bar = tk.Frame(root, bg=BG_LIGHT, height=20)
        status_bar.pack(fill="x", side="bottom", padx=4, pady=(0,4))
        status_bar.pack_propagate(False)

        self.status_label = tk.Label(
            status_bar, text="Bereit | Portfolio_Syskomp_pA.xlsx",
            bg=BG_LIGHT, fg=TEXT_GRAY, font=("Segoe UI", 8), anchor="w"
        )
        self.status_label.pack(side="left", fill="x", expand=True, padx=4)

        # Auto-load data
        self.load_data()

    def load_data(self):
        filepath = "Portfolio_Syskomp_pA.xlsx"
        if not os.path.exists(filepath):
            return

        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            self.data.clear()

            for row_idx in range(2, ws.max_row + 1):
                row_dict = {}
                for col_idx, col_letter in enumerate(['A','B','C','D','E','F','G','H'], start=1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_dict[col_letter] = str(cell_value).strip() if cell_value else ""

                for col_letter in ['A','B','D','E','F','H']:
                    value = row_dict.get(col_letter, "")
                    if value:
                        self.data[col_letter][value] = row_dict

            self.status_label.config(text=f"Bereit | {ws.max_row-1} Zeilen geladen")
        except:
            pass

    def get_col_letter(self, selection):
        return selection.split(':')[0].strip()

    def validate_conversion(self, from_col, to_col, mode):
        if from_col not in ['A','B'] and to_col not in ['A','B']:
            return False, "❌ A oder B muss involviert sein"
        if mode == "extern" and to_col not in ['A','B']:
            return False, "❌ Extern: Nur → A/B erlaubt"
        return True, ""

    def convert_single(self):
        from_col = self.get_col_letter(self.from_var.get())
        to_col = self.get_col_letter(self.to_var.get())
        search_value = self.search_entry.get().strip()
        mode = self.mode_var.get()

        self.result_text.delete(1.0, tk.END)
        self.image_label.config(image='', text='')

        if not search_value:
            self.result_text.insert(1.0, "⚠ Bitte Nummer eingeben")
            return

        valid, error = self.validate_conversion(from_col, to_col, mode)
        if not valid:
            self.result_text.insert(1.0, error)
            return

        if mode == "extern":
            self.request_count += 1
            if self.request_count >= 50:
                import time
                time.sleep(2)

        row_data = self.data.get(from_col, {}).get(search_value)
        if not row_data:
            self.result_text.insert(1.0, f"✗ '{search_value}' nicht gefunden")
            self.status_label.config(text="Nicht gefunden")
            return

        if to_col in ['F','G']:
            result = f"{row_data.get('F','-')} / {row_data.get('G','-')}"
        else:
            result = row_data.get(to_col, '-')

        desc = row_data.get('C','').replace(';','\n')

        self.result_text.insert(1.0, f"✓ Conversion erfolgreich\n\n{search_value} → {result}\n\nBeschreibung:\n{desc}")
        self.status_label.config(text=f"{from_col} → {to_col}")

        # Show image
        ask_nr = row_data.get('H','')
        alvaris_nr = row_data.get('F','')
        bosch_nr = row_data.get('E','')
        item_nr = row_data.get('D','')

        if alvaris_nr and alvaris_nr != '-':
            self.show_image(alvaris_nr, 'alvaris')
        elif ask_nr and ask_nr != '-':
            if bosch_nr and len(bosch_nr) == 10:
                self.show_image(ask_nr, 'bosch')
            elif item_nr and '.' in item_nr:
                self.show_image(ask_nr, 'item')

    def show_image(self, artnr, source_type):
        self.image_label.config(image='', text='')
        self.current_image = None

        if not artnr or artnr == '-':
            return

        script_dir = os.path.dirname(os.path.abspath(__file__))
        possible_dirs = []

        if source_type == 'bosch':
            possible_dirs = [
                os.path.join(script_dir, "ASK_CATALOG", "ASKbosch-all-images"),
            ]
        elif source_type == 'item':
            possible_dirs = [
                os.path.join(script_dir, "ASK_CATALOG", "ASKitem-all-images"),
            ]
        elif source_type == 'alvaris':
            possible_dirs = [
                os.path.join(script_dir, "ALVARIS_CATALOG", "alvaris-all-images"),
            ]

        img_path = None
        for img_dir in possible_dirs:
            test_path = os.path.join(img_dir, f"{artnr}.png")
            if os.path.exists(test_path):
                img_path = test_path
                break

        if not img_path:
            return

        try:
            img = Image.open(img_path)

            if source_type == 'alvaris':
                width, height = img.size
                crop_height = int(height * 0.7)
                img = img.crop((0, 0, width, crop_height))

            max_width, max_height = 250, 220
            width, height = img.size
            ratio = min(max_width / width, max_height / height)
            if ratio < 1:
                img = img.resize((int(width * ratio), int(height * ratio)), Image.Resampling.LANCZOS)

            photo = ImageTk.PhotoImage(img)
            self.current_image = photo
            self.image_label.config(image=photo)
        except:
            pass

    def load_batch_file(self):
        filepath = filedialog.askopenfilename(
            title="Batch-Datei",
            filetypes=[("Text/CSV", "*.txt;*.csv"), ("Alle", "*.*")]
        )
        if filepath:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
                self.batch_text.delete(1.0, tk.END)
                self.batch_text.insert(1.0, content)
                self.file_label.config(text=os.path.basename(filepath))
            except Exception as e:
                messagebox.showerror("Fehler", str(e))

    def process_batch(self):
        input_text = self.batch_text.get(1.0, tk.END).strip()
        if not input_text:
            messagebox.showwarning("Warnung", "Keine Daten")
            return

        lines = [l.strip() for l in input_text.split('\n') if l.strip()]
        target_col = self.batch_target_var.get().split(':')[0].strip()

        self.batch_text.delete(1.0, tk.END)
        self.batch_text.insert(1.0, "Batch Ergebnisse:\n" + "="*50 + "\n\n")

        found, not_found = 0, 0
        for search_value in lines:
            row_data = None
            for col in ['A','B','D','E','F','H']:
                if search_value in self.data.get(col, {}):
                    row_data = self.data[col][search_value]
                    break

            if row_data:
                result = row_data.get(target_col, '-')
                desc = row_data.get('C','').replace(';',' | ')
                self.batch_text.insert(tk.END, f"✓ {search_value} → {result}\n  {desc}\n\n")
                found += 1
            else:
                self.batch_text.insert(tk.END, f"✗ {search_value} → Nicht gefunden\n\n")
                not_found += 1

        self.batch_text.insert(tk.END, "="*50 + f"\n{found} gefunden, {not_found} nicht gefunden\n")
        self.status_label.config(text=f"Batch: {found}/{len(lines)} gefunden")

    def save_batch(self):
        content = self.batch_text.get(1.0, tk.END).strip()
        if not content:
            messagebox.showwarning("Warnung", "Nichts zum Speichern")
            return

        filepath = filedialog.asksaveasfilename(
            title="Speichern",
            defaultextension=".txt",
            filetypes=[("Text", "*.txt"), ("CSV", "*.csv"), ("Alle", "*.*")]
        )
        if filepath:
            try:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(content)
                messagebox.showinfo("Erfolg", "Gespeichert")
            except Exception as e:
                messagebox.showerror("Fehler", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = ConversionApp(root)
    root.mainloop()
