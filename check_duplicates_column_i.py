import openpyxl
from collections import Counter

# Load workbook
file_path = "Portfolio_Syskomp_pA_neu.xlsx"
print(f"Lade Excel-Datei: {file_path}")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"Worksheet: {ws.title}")
print(f"Max Row: {ws.max_row}")

# Collect all values from column I
column_i_values = []
row_value_map = {}  # Map value to list of row numbers

for row_idx in range(1, ws.max_row + 1):
    cell_i = ws.cell(row=row_idx, column=9)  # Column I
    if cell_i.value and str(cell_i.value).strip():
        value = str(cell_i.value).strip()
        column_i_values.append(value)

        if value not in row_value_map:
            row_value_map[value] = []
        row_value_map[value].append(row_idx)

# Count occurrences
value_counts = Counter(column_i_values)

print(f"\nGesamt {len(column_i_values)} Werte in Spalte I")
print(f"Davon {len(value_counts)} eindeutige Werte")

# Find duplicates
duplicates = {value: count for value, count in value_counts.items() if count > 1}

if duplicates:
    print(f"\nDUPLIKATE GEFUNDEN: {len(duplicates)} Werte kommen mehrfach vor:\n")

    for value, count in sorted(duplicates.items(), key=lambda x: x[1], reverse=True):
        rows = row_value_map[value]
        print(f"  '{value}' kommt {count}x vor in Zeilen: {rows}")
else:
    print("\nKeine Duplikate gefunden - alle Werte in Spalte I sind eindeutig!")
