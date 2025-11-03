import openpyxl

file_path = "Portfolio_Syskomp_pA_neu.xlsx"

# Load with data_only=True
print("=== Mit data_only=True (berechnete Werte) ===")
wb_data = openpyxl.load_workbook(file_path, data_only=True)
ws_data = wb_data.active

row_num = 37
print(f"Zeile {row_num}, Spalte I: '{ws_data.cell(row=row_num, column=9).value}'")
print(f"Zeile {row_num}, Spalte L: '{ws_data.cell(row=row_num, column=12).value}'")

print("\n=== Ohne data_only (Formeln) ===")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"Zeile {row_num}, Spalte I: '{ws.cell(row=row_num, column=9).value}'")
print(f"Zeile {row_num}, Spalte L: '{ws.cell(row=row_num, column=12).value}'")

print("\n=== Alle Zellen in Zeile 37 (mit berechneten Werten) ===")
for col_idx in range(1, 27):
    cell = ws_data.cell(row=row_num, column=col_idx)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"  {col_letter}: {cell.value}")
