"""
Flask backend API for Portfolio conversion tool
Adapted for Vercel serverless deployment
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from openpyxl import load_workbook
import os
from collections import defaultdict

app = Flask(__name__)
CORS(app)

# Configure rate limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Data storage
data = defaultdict(dict)
COLUMN_NAMES = {
    'A': 'Syskomp neu',
    'B': 'Syskomp alt',
    'C': 'Beschreibung',
    'D': 'Item',
    'E': 'Bosch',
    'F': 'Alvaris Artnr',
    'G': 'Alvaris Matnr',
    'H': 'ASK'
}

def load_data():
    """Load Portfolio_Syskomp_pA.xlsx data"""
    # Try multiple possible paths for Vercel deployment
    possible_paths = [
        os.path.join(os.path.dirname(__file__), '..', 'Portfolio_Syskomp_pA.xlsx'),
        os.path.join(os.getcwd(), 'Portfolio_Syskomp_pA.xlsx'),
        'Portfolio_Syskomp_pA.xlsx'
    ]

    filepath = None
    for path in possible_paths:
        if os.path.exists(path):
            filepath = path
            break

    if not filepath:
        print(f"ERROR: Portfolio_Syskomp_pA.xlsx not found in: {possible_paths}")
        return False

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        data.clear()

        for row_idx in range(2, ws.max_row + 1):
            row_dict = {}
            for col_idx, col_letter in enumerate(['A','B','C','D','E','F','G','H'], start=1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_dict[col_letter] = str(cell_value).strip() if cell_value else ""

            # Index all searchable columns (all except C which is description)
            for col_letter in ['A','B','D','E','F','G','H']:
                value = row_dict.get(col_letter, "")
                if value and value != "None":
                    data[col_letter][value] = row_dict

        print(f"Data loaded: {ws.max_row-1} rows, {sum(len(v) for v in data.values())} indexed entries")
        return True
    except Exception as e:
        print(f"ERROR loading data: {e}")
        return False

def validate_conversion(from_col, to_col, mode):
    """Validate conversion rules based on mode"""
    # Rule: A or B must be involved
    if from_col not in ['A','B'] and to_col not in ['A','B']:
        return False, "Konvertierung muss A oder B beinhalten"

    # External mode (Vercel): only allow conversions TO A or B
    if mode == "extern" and to_col not in ['A','B']:
        return False, "Extern-Modus: Nur Konvertierung nach A oder B erlaubt"

    # Internal mode: FROM must be A or B
    if mode == "intern" and from_col not in ['A','B']:
        return False, "Intern-Modus: Konvertierung muss von A oder B starten"

    return True, ""

@app.route('/api/search', methods=['POST'])
@limiter.limit("30 per minute")
def search_all():
    """Search in all columns and return all matches"""
    try:
        req_data = request.json
        search_value = req_data.get('number', '').strip()

        if not search_value:
            return jsonify({'error': 'Keine Nummer angegeben'}), 400

        # Search in all columns
        matches = []
        seen_rows = set()  # Track unique rows to avoid duplicates

        for col_letter in ['A','B','D','E','F','G','H']:
            row_data = data.get(col_letter, {}).get(search_value)

            if row_data:
                # Create a unique key for this row (using Syskomp A and B)
                row_key = (row_data.get('A', ''), row_data.get('B', ''))

                if row_key not in seen_rows:
                    seen_rows.add(row_key)

                    # Get description
                    description = row_data.get('C', '').replace(';', '\n')

                    # Image info (for local mode, not Vercel)
                    image_info = None
                    alvaris_nr = row_data.get('F', '')
                    ask_nr = row_data.get('H', '')

                    if alvaris_nr and alvaris_nr != '-' and alvaris_nr != 'None':
                        image_info = {
                            'type': 'alvaris',
                            'artnr': alvaris_nr,
                            'crop_top_70': True
                        }
                    elif ask_nr and ask_nr != '-' and ask_nr != 'None':
                        image_info = {
                            'type': 'ask',
                            'artnr': ask_nr,
                            'crop_top_70': False
                        }

                    matches.append({
                        'found_in_col': col_letter,
                        'found_in_col_name': COLUMN_NAMES.get(col_letter, col_letter),
                        'syskomp_neu': row_data.get('A', '-'),
                        'syskomp_alt': row_data.get('B', '-'),
                        'item': row_data.get('D', '-'),
                        'bosch': row_data.get('E', '-'),
                        'alvaris_artnr': row_data.get('F', '-'),
                        'alvaris_matnr': row_data.get('G', '-'),
                        'ask': row_data.get('H', '-'),
                        'description': description,
                        'image': image_info
                    })

        if not matches:
            return jsonify({
                'found': False,
                'search_term': search_value
            })

        return jsonify({
            'found': True,
            'search_term': search_value,
            'count': len(matches),
            'matches': matches
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/convert', methods=['POST'])
@limiter.limit("30 per minute")
def convert_single():
    """Single number conversion"""
    try:
        req_data = request.json
        from_col = req_data.get('from_col', '').upper()
        to_col = req_data.get('to_col', '').upper()
        search_value = req_data.get('number', '').strip()
        mode = req_data.get('mode', 'extern')  # Default to extern for Vercel

        if not search_value:
            return jsonify({'error': 'Keine Nummer angegeben'}), 400

        # Validate conversion
        valid, error = validate_conversion(from_col, to_col, mode)
        if not valid:
            return jsonify({'error': error}), 400

        # Search for number
        row_data = data.get(from_col, {}).get(search_value)

        if not row_data:
            return jsonify({
                'found': False,
                'search_term': search_value,
                'from_col': from_col,
                'to_col': to_col
            })

        # Get result
        if to_col in ['F', 'G']:
            # Special case for Alvaris: show both Artnr and Matnr
            result_value = f"{row_data.get('F', '-')} / {row_data.get('G', '-')}"
        else:
            result_value = row_data.get(to_col, '-')

        # Get description and convert semicolons to newlines
        description = row_data.get('C', '').replace(';', '\n')

        # Image info (path checking happens on client side for Vercel)
        image_info = None
        alvaris_nr = row_data.get('F', '')
        ask_nr = row_data.get('H', '')

        if alvaris_nr and alvaris_nr != '-' and alvaris_nr != 'None':
            image_info = {
                'type': 'alvaris',
                'artnr': alvaris_nr,
                'crop_top_70': True
            }
        elif ask_nr and ask_nr != '-' and ask_nr != 'None':
            image_info = {
                'type': 'ask',
                'artnr': ask_nr,
                'crop_top_70': False
            }

        return jsonify({
            'found': True,
            'from_col': from_col,
            'from_col_name': COLUMN_NAMES.get(from_col, from_col),
            'to_col': to_col,
            'to_col_name': COLUMN_NAMES.get(to_col, to_col),
            'search_value': search_value,
            'result_value': result_value,
            'description': description,
            'row_data': row_data,
            'image': image_info
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-convert', methods=['POST'])
@limiter.limit("10 per minute")
def batch_convert():
    """Batch conversion"""
    try:
        req_data = request.json
        numbers = req_data.get('numbers', [])
        target_col = req_data.get('target_col', 'A').upper()
        mode = req_data.get('mode', 'extern')  # Default to extern for Vercel

        if target_col not in ['A', 'B']:
            return jsonify({'error': 'Batch-Konvertierung nur nach A oder B erlaubt'}), 400

        results = []

        for idx, search_value in enumerate(numbers):
            search_value = str(search_value).strip()

            if not search_value or search_value == 'None':
                results.append({
                    'index': idx,
                    'input': search_value,
                    'output': None,
                    'status': 'empty'
                })
                continue

            # Search in all columns
            row_data = None
            found_in_col = None

            for col in ['A','B','D','E','F','G','H']:
                if search_value in data.get(col, {}):
                    row_data = data[col][search_value]
                    found_in_col = col
                    break

            if row_data:
                # Validate conversion
                valid, error = validate_conversion(found_in_col, target_col, mode)

                if valid:
                    result_value = row_data.get(target_col, None)
                    if result_value and result_value != 'None':
                        results.append({
                            'index': idx,
                            'input': search_value,
                            'output': result_value,
                            'status': 'success',
                            'from_col': found_in_col
                        })
                    else:
                        results.append({
                            'index': idx,
                            'input': search_value,
                            'output': None,
                            'status': 'not_found'
                        })
                else:
                    results.append({
                        'index': idx,
                        'input': search_value,
                        'output': None,
                        'status': 'invalid_conversion',
                        'message': error
                    })
            else:
                results.append({
                    'index': idx,
                    'input': search_value,
                    'output': None,
                    'status': 'not_found'
                })

        success_count = sum(1 for r in results if r['status'] == 'success')

        return jsonify({
            'total': len(numbers),
            'success': success_count,
            'failed': len(numbers) - success_count,
            'results': results
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({
        'status': 'ok',
        'data_loaded': len(data) > 0,
        'rows_indexed': sum(len(v) for v in data.values()),
        'columns': list(COLUMN_NAMES.keys())
    })

@app.route('/api/columns', methods=['GET'])
def get_columns():
    """Get available columns"""
    return jsonify({
        'columns': COLUMN_NAMES
    })

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Get statistics about the data"""
    return jsonify({
        'syskomp': len(data.get('A', {})),  # Count unique Syskomp neu entries
        'item': len(data.get('D', {})),
        'bosch': len(data.get('E', {})),
        'alvaris': len(data.get('F', {})),
        'ask': len(data.get('H', {}))
    })

# Load data on module initialization
print("Initializing Portfolio Conversion API...")
load_data()

# For Vercel serverless
app = app
