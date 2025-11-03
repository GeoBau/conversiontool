from openpyxl import load_workbook
from collections import defaultdict

# Replicate conversion_app.py data loading
data = defaultdict(dict)

wb = load_workbook('Portfolio_Syskomp_pA.xlsx', data_only=True)
ws = wb.active

for row_idx in range(2, ws.max_row + 1):
    row_dict = {}
    for col_idx, col_letter in enumerate(['A','B','C','D','E','F','G','H'], start=1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_dict[col_letter] = str(cell_value).strip() if cell_value else ""

    for col_letter in ['A','B','D','E','F','H']:
        value = row_dict.get(col_letter, "")
        if value:
            data[col_letter][value] = row_dict

# Test conversion
search_value = '402150012'
from_col = 'B'  # Syskomp alt
to_col = 'A'    # Syskomp neu

print(f"Searching for: {search_value}")
print(f"From column: {from_col} (Syskomp alt)")
print(f"To column: {to_col} (Syskomp neu)")
print()

row_data = data.get(from_col, {}).get(search_value)
if row_data:
    result = row_data.get(to_col, '-')
    desc = row_data.get('C', '')
    print(f"FOUND!")
    print(f"Result: {search_value} -> {result}")
    print(f"Description: {desc}")
else:
    print(f"NOT FOUND in data['{from_col}']")
    print()
    print(f"Available keys in data: {list(data.keys())}")
    print(f"Number of entries in column B: {len(data.get('B', {}))}")
    print(f"Sample B keys: {list(data.get('B', {}).keys())[:10]}")
