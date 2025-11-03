import csv
from openpyxl import load_workbook

def is_syskomp_number(nr):
    """Check if number is Syskomp format: 9-digit, starts with 2 or 4"""
    if not nr:
        return False
    nr_str = str(nr).strip()
    return (nr_str.isdigit() and
            len(nr_str) == 9 and
            nr_str[0] in ['2', '4'])

def is_bosch_number(nr):
    """Check if number is Bosch format: 10-digit"""
    if not nr:
        return False
    nr_str = str(nr).strip()
    return nr_str.isdigit() and len(nr_str) == 10

# Step 1: Load ArtNrn.csv and create mapping Syskomp -> Bosch
print("Lade ArtNrn.csv...")
syskomp_to_bosch = {}

artnrn_path = r"\\sys-ts19-1\c$\ArtNrConverter\ArtNrn.csv"
# Fallback to local if network not available
import os
if not os.path.exists(artnrn_path):
    artnrn_path = "Vorlagen/ArtNrn.csv"

with open(artnrn_path, 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)  # Skip header

    for row in reader:
        if len(row) < 2:
            continue

        col_a = str(row[0]).strip() if row[0] else ""
        col_b = str(row[1]).strip() if row[1] else ""

        # Determine which is Syskomp and which is Bosch
        syskomp_nr = None
        bosch_nr = None

        if is_syskomp_number(col_a) and is_bosch_number(col_b):
            syskomp_nr = col_a
            bosch_nr = col_b
        elif is_bosch_number(col_a) and is_syskomp_number(col_b):
            syskomp_nr = col_b
            bosch_nr = col_a

        if syskomp_nr and bosch_nr:
            syskomp_to_bosch[syskomp_nr] = bosch_nr

print(f"Gefunden: {len(syskomp_to_bosch)} Syskomp <-> Bosch Zuordnungen")

# Step 2: Load Excel and add Bosch numbers to column F
excel_path = "Portfolio_Syskomp_pA.xlsx"
print(f"\nLade Excel-Datei: {excel_path}")

wb = load_workbook(excel_path)
ws = wb.active

print(f"Worksheet: {ws.title}")
print(f"Max Row: {ws.max_row}")

matches_found = 0
no_match_count = 0

# Process each row
for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
    # Column B (index 2) = Syskomp number
    syskomp_cell = ws.cell(row=row_idx, column=2)
    syskomp_value = str(syskomp_cell.value).strip() if syskomp_cell.value else ""

    # Only process if column B has a Syskomp number
    if is_syskomp_number(syskomp_value):
        # Look up Bosch number
        bosch_nr = syskomp_to_bosch.get(syskomp_value)

        if bosch_nr:
            # Write to column F (index 6)
            ws.cell(row=row_idx, column=6, value=bosch_nr)
            matches_found += 1
            if matches_found <= 5:  # Show first 5 matches
                print(f"  Zeile {row_idx}: {syskomp_value} -> {bosch_nr}")
        else:
            no_match_count += 1

print(f"\nErgebnis:")
print(f"  {matches_found} Bosch-Nummern in Spalte F eingefügt")
print(f"  {no_match_count} Syskomp-Nummern ohne passende Bosch-Nummer")

# Save Excel
wb.save(excel_path)
print(f"\nDatei gespeichert: {excel_path}")
